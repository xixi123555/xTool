from __future__ import annotations
import hashlib
import base64
import csv
import json
import math
import os
import re
import zipfile
import xml.etree.ElementTree as ET
from urllib import request as urllib_request
from typing import Any, Dict, List, Optional, Tuple
import openpyxl
import xlrd
from app.core.settings import settings
from app.repositories.vector_store_repository import (
    delete_documents_by_source_prefix,
    ensure_vector_tables,
    get_document_by_source,
    list_candidates,
    list_chat_file_documents,
    search_candidates_by_keywords,
    upsert_document,
    upsert_vector,
)
from app.schemas.chat import RagSource


_TOKEN_RE = re.compile(r"[a-zA-Z0-9\u4e00-\u9fff]+")
_DIM = 128
_MAX_EXTRACT_LINES = 20000
_EMBEDDING_MODEL: Any = None
_TERM_SYNONYMS: Dict[str, List[str]] = {
    "整装总计": ["整装总计", "整装价格", "整装费用", "整装金额", "整装多少钱", "整装报价"],
    "半包": ["半包", "半包价格", "半包费用", "半包金额", "半包多少钱", "半包报价"],
    "设计费": ["设计费", "设计费用", "设计价格", "设计金额", "设计多少钱", "设计报价"],
    "主材": ["主材", "主材费用", "主材价格", "主材金额"],
    "全房定制": ["全房定制", "全屋定制", "全房定制费用", "全屋定制费用"],
    "灯具照明": ["灯具照明", "灯具", "照明", "灯具费用", "照明费用"],
}
_ALIAS_TO_CANONICAL: Dict[str, str] = {}
for _canonical, _aliases in _TERM_SYNONYMS.items():
    for _alias in _aliases:
        _ALIAS_TO_CANONICAL[_alias] = _canonical


def _excel_file_key(file_url: str) -> str:
    return hashlib.md5((file_url or "").encode("utf-8")).hexdigest()


def _tokenize(text: str) -> List[str]:
    return [t.lower() for t in _TOKEN_RE.findall(text or "")]


def _extract_semantic_canonical_terms(text: str) -> List[str]:
    raw = str(text or "")
    found: List[str] = []
    for alias, canonical in _ALIAS_TO_CANONICAL.items():
        if alias and alias in raw and canonical not in found:
            found.append(canonical)
    return found


def _expand_query_terms(query: str) -> List[str]:
    q = str(query or "")
    expanded: List[str] = []
    for tk in _tokenize(q):
        if len(tk) >= 2:
            expanded.append(tk)
    canonicals = _extract_semantic_canonical_terms(q)
    for canonical in canonicals:
        expanded.append(canonical)
        for alias in _TERM_SYNONYMS.get(canonical, []):
            expanded.append(alias)
    # 保序去重
    return list(dict.fromkeys([x for x in expanded if x]))


def _fallback_hash_embed(text: str) -> List[float]:
    vec = [0.0] * _DIM
    tokens = _tokenize(text)
    if not tokens:
        return vec
    for token in tokens:
        hv = int(hashlib.md5(token.encode("utf-8")).hexdigest(), 16)
        idx = hv % _DIM
        vec[idx] += 1.0
    norm = math.sqrt(sum(v * v for v in vec)) or 1.0
    return [v / norm for v in vec]


def _get_embedding_model() -> Optional[Any]:
    global _EMBEDDING_MODEL
    if _EMBEDDING_MODEL is not None:
        return _EMBEDDING_MODEL
    try:
        from sentence_transformers import SentenceTransformer  # type: ignore
        _EMBEDDING_MODEL = SentenceTransformer(settings.embedding_model_name)
        return _EMBEDDING_MODEL
    except Exception:
        _EMBEDDING_MODEL = None
        return None


def _embed_text(text: str) -> List[float]:
    """
    主向量化：BAAI/bge-small-zh-v1.5
    兜底：hash 向量（避免依赖未就绪导致流程中断）
    """
    content = (text or "").strip()
    if not content:
        return [0.0] * _DIM

    model = _get_embedding_model()
    if model is None:
        return _fallback_hash_embed(content)
    try:
        vec = model.encode(
            [content],
            normalize_embeddings=True,
            convert_to_numpy=True,
        )[0]
        return [float(x) for x in vec.tolist()]
    except Exception:
        return _fallback_hash_embed(content)


def _cosine(a: List[float], b: List[float]) -> float:
    if not a or not b or len(a) != len(b):
        return 0.0
    return float(sum(x * y for x, y in zip(a, b)))


def _keyword_overlap(query: str, content: str) -> float:
    q = set(_tokenize(query))
    c = set(_tokenize(content))
    if not q or not c:
        return 0.0
    return len(q.intersection(c)) / max(len(q), 1)


def _hybrid_score(query: str, query_vec: List[float], doc_text: str, doc_vec: List[float]) -> float:
    dense_score = _cosine(query_vec, doc_vec)
    sparse_score = _keyword_overlap(query, doc_text)
    return dense_score * 0.65 + sparse_score * 0.35


def _query_entities(query: str) -> Tuple[str, str]:
    q = (query or "").strip()
    date_tokens = re.findall(r"\d{4}年\d{1,2}月\d{1,2}日", q)
    cn_tokens = re.findall(r"[\u4e00-\u9fa5]{2,4}", q)
    ignore_parts = ("什么", "做了", "哪些", "哪天", "当天", "内容", "事项", "文档", "做啥", "做过")
    raw_name_tokens = [t for t in cn_tokens if not any(part in t for part in ignore_parts)]
    stop_chars = ("在", "于", "把", "将")
    name_tokens: List[str] = []
    for t in raw_name_tokens:
        cleaned = t.strip()
        while len(cleaned) > 2 and cleaned.endswith(stop_chars):
            cleaned = cleaned[:-1]
        if len(cleaned) >= 2:
            name_tokens.append(cleaned)
    target_name = name_tokens[0] if name_tokens else ""
    target_date = date_tokens[0] if date_tokens else ""
    return target_name, target_date


def _resolve_entities_for_query(
    query: str,
    slot_person: Optional[str] = None,
    slot_date_cn: Optional[str] = None,
) -> Tuple[str, str]:
    """优先使用 LLM 槽位，否则回退到规则从 query 解析。"""
    name = (slot_person or "").strip()
    date_cn = (slot_date_cn or "").strip()
    if not name or not date_cn:
        n2, d2 = _query_entities(query)
        if not name:
            name = n2
        if not date_cn:
            date_cn = d2
    return name, date_cn


def _date_alternatives(date_text: str) -> List[str]:
    d = (date_text or "").strip()
    if not d:
        return []
    m = re.match(r"(\d{4})年(\d{1,2})月(\d{1,2})日", d)
    if not m:
        return [d]
    y, mo, da = m.group(1), int(m.group(2)), int(m.group(3))
    return [d, f"{y}/{mo}/{da}", f"{y}/{mo:02d}/{da:02d}", f"{y}-{mo:02d}-{da:02d}"]


def _best_context_snippet(
    text: str,
    query: str,
    remain: int,
    entity_name: Optional[str] = None,
    entity_date: Optional[str] = None,
) -> str:
    content = (text or "").strip()
    if len(content) <= remain:
        return content
    if entity_name is not None or entity_date is not None:
        target_name = (entity_name or "").strip()
        target_date = (entity_date or "").strip()
    else:
        target_name, target_date = _query_entities(query)
    query_tokens = [t for t in _tokenize(query) if t]
    date_tokens = _date_alternatives(target_date)
    lines = [ln for ln in content.splitlines() if ln.strip()]
    if not lines:
        return content[:remain]

    best_idx = 0
    best_score = -1
    for idx, line in enumerate(lines):
        score = 0
        if target_name and target_name in line:
            score += 4
        if date_tokens and any(dt in line for dt in date_tokens):
            score += 4
        for tk in query_tokens[:8]:
            if tk and tk in line:
                score += 1
        if score > best_score:
            best_score = score
            best_idx = idx

    if best_score <= 0:
        return content[:remain]

    window_lines: List[str] = []
    consumed = 0
    start = max(0, best_idx - 30)
    end = min(len(lines), best_idx + 31)
    for ln in lines[start:end]:
        n = len(ln) + 1
        if consumed + n > remain:
            break
        window_lines.append(ln)
        consumed += n
    return "\n".join(window_lines) if window_lines else content[:remain]


def _source_weight(source_type: str) -> float:
    if source_type in ("chat_excel_row",):
        return 1.35
    if source_type in ("chat_file", "manual_text"):
        return 1.2
    if source_type == "chat_message":
        return 0.9
    return 1.0


async def ensure_ready() -> None:
    await ensure_vector_tables()


async def upsert_knowledge_document(
    room_id: str,
    source_type: str,
    source_id: Optional[str],
    content_text: str,
    metadata: Optional[Dict] = None,
) -> int:
    doc_id = await upsert_document(
        room_id=room_id,
        source_type=source_type,
        source_id=source_id,
        content_text=content_text,
        metadata_json=json.dumps(metadata or {}, ensure_ascii=False),
    )
    if doc_id <= 0:
        return 0
    vector = _embed_text(content_text)
    await upsert_vector(doc_id=doc_id, vector_json=json.dumps(vector))
    return doc_id


async def ingest_chat_message(
    message_id: int,
    room_id: str,
    text: str,
    extra: Optional[Dict] = None,
) -> int:
    if not text.strip():
        return 0
    return await upsert_knowledge_document(
        room_id=room_id,
        source_type="chat_message",
        source_id=str(message_id),
        content_text=text.strip(),
        metadata=extra or {},
    )


async def ingest_uploaded_file(
    room_id: str,
    file_url: str,
    file_name: str,
    mime_type: str,
    file_size: int,
    local_path: Optional[str] = None,
) -> int:
    lower_name = (file_name or "").lower()
    lower_mime = (mime_type or "").lower()
    is_excel = (
        lower_name.endswith((".xlsx", ".xls"))
        or "ms-excel" in lower_mime
        or "spreadsheetml" in lower_mime
    )
    if is_excel and local_path and os.path.exists(local_path):
        return await _ingest_excel_rows(
            room_id=room_id,
            file_url=file_url,
            file_name=file_name,
            mime_type=mime_type,
            file_size=file_size,
            local_path=local_path,
        )

    extracted_text = _extract_file_text(local_path, file_name, mime_type)
    summary = (
        f"文件名: {file_name}\n类型: {mime_type}\n大小: {file_size}\n链接: {file_url}\n\n"
        f"抽取内容:\n{extracted_text if extracted_text else '（暂未抽取到可读文本）'}"
    )
    return await upsert_knowledge_document(
        room_id=room_id,
        source_type="chat_file",
        source_id=file_url,
        content_text=summary,
        metadata={"name": file_name, "mime_type": mime_type, "size": file_size, "url": file_url},
    )


def _file_sha256(path: str) -> str:
    sha = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            block = f.read(1024 * 1024)
            if not block:
                break
            sha.update(block)
    return sha.hexdigest()


def _normalize_header(value: Any, idx: int) -> str:
    text = str(value).strip() if value is not None else ""
    return text or f"列{idx}"


def _normalize_cell(value: Any) -> str:
    if value is None:
        return "空"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    text = str(value).strip()
    return text if text else "空"


def _excel_row_documents(file_name: str, file_url: str, path: str, mime_type: str = "") -> List[Dict[str, Any]]:
    lower_name = (file_name or "").lower()
    lower_mime = (mime_type or "").lower()
    if lower_name.endswith(".xls") or "ms-excel" in lower_mime:
        return _excel_row_documents_xls(file_name=file_name, file_url=file_url, path=path)
    return _excel_row_documents_xlsx(file_name=file_name, file_url=file_url, path=path)


def _excel_row_documents_xlsx(file_name: str, file_url: str, path: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    docs: List[Dict[str, Any]] = []
    file_key = _excel_file_key(file_url)
    try:
        for ws in wb.worksheets:
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                continue
            headers = [_normalize_header(v, i + 1) for i, v in enumerate(header_row)]
            sheet_key = hashlib.md5(ws.title.encode("utf-8")).hexdigest()[:8]
            for row_idx, row in enumerate(rows_iter, start=2):
                row_values = [_normalize_cell(v) for v in row]
                if all(v == "空" for v in row_values):
                    continue
                pairs: List[str] = []
                row_dict: Dict[str, str] = {}
                for i, head in enumerate(headers):
                    val = row_values[i] if i < len(row_values) else "空"
                    pairs.append(f"{head}:{val}")
                    row_dict[head] = val
                row_text = (
                    f"文件:{file_name}；Sheet:{ws.title}；行号:{row_idx}；"
                    + "；".join(pairs)
                )
                semantic_terms = _extract_semantic_canonical_terms(row_text)
                if semantic_terms:
                    row_text += f"；语义标签:{'、'.join(semantic_terms)}"
                docs.append(
                    {
                        "source_id": f"excelrow:{file_key}:{sheet_key}:{row_idx}",
                        "content_text": row_text,
                        "metadata": {
                            "type": "excel_row",
                            "file_key": file_key,
                            "file_name": file_name,
                            "file_url": file_url,
                            "sheet_name": ws.title,
                            "row_index": row_idx,
                            "row_data": row_dict,
                            "semantic_terms": semantic_terms,
                        },
                    }
                )
    finally:
        wb.close()
    return docs


def _excel_row_documents_xls(file_name: str, file_url: str, path: str) -> List[Dict[str, Any]]:
    with open(path, "rb") as f:
        content = f.read()
    wb = xlrd.open_workbook(file_contents=content)
    docs: List[Dict[str, Any]] = []
    file_key = _excel_file_key(file_url)
    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        if ws.nrows <= 0:
            continue
        header_row = ws.row_values(0)
        headers = [_normalize_header(v, i + 1) for i, v in enumerate(header_row)]
        sheet_key = hashlib.md5(ws.name.encode("utf-8")).hexdigest()[:8]
        for row_idx in range(1, ws.nrows):
            row = ws.row_values(row_idx)
            row_values = [_normalize_cell(v) for v in row]
            if all(v == "空" for v in row_values):
                continue
            pairs: List[str] = []
            row_dict: Dict[str, str] = {}
            max_len = max(len(headers), len(row_values))
            for i in range(max_len):
                head = headers[i] if i < len(headers) else f"列{i + 1}"
                val = row_values[i] if i < len(row_values) else "空"
                pairs.append(f"{head}:{val}")
                row_dict[head] = val
            real_row = row_idx + 1
            row_text = (
                f"文件:{file_name}；Sheet:{ws.name}；行号:{real_row}；"
                + "；".join(pairs)
            )
            semantic_terms = _extract_semantic_canonical_terms(row_text)
            if semantic_terms:
                row_text += f"；语义标签:{'、'.join(semantic_terms)}"
            docs.append(
                {
                    "source_id": f"excelrow:{file_key}:{sheet_key}:{real_row}",
                    "content_text": row_text,
                    "metadata": {
                        "type": "excel_row",
                        "file_key": file_key,
                        "file_name": file_name,
                        "file_url": file_url,
                        "sheet_name": ws.name,
                        "row_index": real_row,
                        "row_data": row_dict,
                        "semantic_terms": semantic_terms,
                    },
                }
            )
    return docs


async def _ingest_excel_rows(
    room_id: str,
    file_url: str,
    file_name: str,
    mime_type: str,
    file_size: int,
    local_path: str,
) -> int:
    file_hash = _file_sha256(local_path)
    file_key = _excel_file_key(file_url)
    manifest_source_id = f"excelmanifest:{file_key}"
    existing_manifest = await get_document_by_source("chat_excel_manifest", manifest_source_id)
    if existing_manifest:
        try:
            meta = json.loads(existing_manifest.get("metadata_json") or "{}")
        except Exception:
            meta = {}
        if meta.get("file_hash") == file_hash:
            return int(existing_manifest["id"])

    rows = _excel_row_documents(file_name=file_name, file_url=file_url, path=local_path, mime_type=mime_type)
    await delete_documents_by_source_prefix("chat_excel_row", f"excelrow:{file_key}:")
    for row in rows:
        await upsert_knowledge_document(
            room_id=room_id,
            source_type="chat_excel_row",
            source_id=row["source_id"],
            content_text=row["content_text"],
            metadata=row["metadata"],
        )

    manifest_text = (
        f"Excel文件:{file_name}\n"
        f"链接:{file_url}\n"
        f"类型:{mime_type}\n"
        f"大小:{file_size}\n"
        f"行文档数:{len(rows)}"
    )
    return await upsert_knowledge_document(
        room_id=room_id,
        source_type="chat_excel_manifest",
        source_id=manifest_source_id,
        content_text=manifest_text,
        metadata={
            "file_hash": file_hash,
            "file_name": file_name,
            "file_url": file_url,
            "mime_type": mime_type,
            "size": file_size,
            "row_count": len(rows),
        },
    )


async def query_knowledge(
    query: str,
    room_id: str,
    top_k: Optional[int] = None,
    candidate_k: Optional[int] = None,
    slot_person: Optional[str] = None,
    slot_date_cn: Optional[str] = None,
) -> Tuple[str, List[RagSource]]:
    q = (query or "").strip()
    if not q:
        return "", []

    q_vec = _embed_text(q)
    target_name, target_date = _resolve_entities_for_query(q, slot_person, slot_date_cn)
    limit = candidate_k or settings.kb_candidate_k
    limit = max(limit, settings.kb_candidate_k_excel)
    candidates = await list_candidates(room_id=room_id, limit=limit)
    # 补充关键词召回：避免仅按 updated_at 截断导致历史 Excel 行文档被遗漏
    expanded_terms = _expand_query_terms(q)
    query_semantic_terms = _extract_semantic_canonical_terms(q)
    if expanded_terms:
        keyword_docs = await search_candidates_by_keywords(
            room_id=room_id,
            keywords=expanded_terms[:12],
            limit=max(limit, 240),
            source_types=("chat_excel_row", "chat_file", "manual_text"),
        )
        if keyword_docs:
            merged: Dict[int, Dict] = {}
            for item in candidates:
                try:
                    merged[int(item.get("id"))] = item
                except Exception:
                    continue
            for item in keyword_docs:
                try:
                    merged[int(item.get("id"))] = item
                except Exception:
                    continue
            candidates = list(merged.values())
    ranked: List[Tuple[float, Dict]] = []
    for item in candidates:
        source_type = str(item.get("source_type") or "")
        text_value = item.get("content_text") or ""
        metadata_raw = item.get("metadata_json") or "{}"
        try:
            metadata = json.loads(metadata_raw) if isinstance(metadata_raw, str) else (metadata_raw or {})
        except Exception:
            metadata = {}
        # 避免历史机器人问答反哺污染检索结果
        if source_type == "chat_message" and "用户问题：" in text_value and "智能体回答：" in text_value:
            continue
        raw_vec = item.get("vector_json") or "[]"
        try:
            doc_vec = json.loads(raw_vec)
        except Exception:
            doc_vec = []
        score = _hybrid_score(q, q_vec, text_value, doc_vec) * _source_weight(source_type)
        date_alts = _date_alternatives(target_date)
        has_name = bool(target_name and target_name in text_value)
        has_date = bool(date_alts and any(dt in text_value for dt in date_alts))
        if has_name:
            score *= 1.2
        if has_date:
            score *= 1.35
        if has_name and has_date:
            score *= 1.5
        # 问题明确带日期时，缺日期的通用手工知识降权，避免挤占上下文窗口
        if source_type == "manual_text" and target_date and not has_date:
            score *= 0.45
        # 文件只有元信息而没有抽取内容时降权，避免压过真实内容文档
        if source_type == "chat_file" and "抽取内容:" not in text_value:
            score *= 0.35
        # 结构化表格行在 query token 命中时额外提权
        if source_type == "chat_excel_row":
            row_overlap = _keyword_overlap(" ".join(expanded_terms) or q, text_value)
            score *= (1.0 + min(0.5, row_overlap))
            if target_name and target_name in text_value:
                score *= 1.1
            source_semantic_terms = metadata.get("semantic_terms") if isinstance(metadata.get("semantic_terms"), list) else []
            if query_semantic_terms and source_semantic_terms:
                overlap = len(set(query_semantic_terms).intersection(set([str(x) for x in source_semantic_terms])))
                if overlap > 0:
                    score *= (1.0 + min(0.45, 0.2 * overlap))
        ranked.append((score, item))
    ranked.sort(key=lambda x: x[0], reverse=True)
    selected = ranked[: max(1, top_k or settings.kb_top_k)]

    snippets: List[str] = []
    sources: List[RagSource] = []
    consumed = 0
    for score, item in selected:
        text = (item.get("content_text") or "").strip()
        if not text:
            continue
        remain = settings.kb_max_context_chars - consumed
        if remain <= 0:
            break
        snippet = _best_context_snippet(text, q, remain, target_name, target_date)
        consumed += len(snippet)
        snippets.append(snippet)
        metadata_raw = item.get("metadata_json") or "{}"
        try:
            source_meta = json.loads(metadata_raw) if isinstance(metadata_raw, str) else (metadata_raw or {})
        except Exception:
            source_meta = {}
        sources.append(
            RagSource(
                doc_id=int(item["id"]),
                source_type=str(item.get("source_type") or "unknown"),
                source_id=str(item.get("source_id")) if item.get("source_id") is not None else None,
                score=round(float(score), 4),
                snippet=snippet[:180],
                sheet_name=str(source_meta.get("sheet_name")) if source_meta.get("sheet_name") else None,
                row_index=int(source_meta.get("row_index")) if source_meta.get("row_index") else None,
                row_data=source_meta.get("row_data") if isinstance(source_meta.get("row_data"), dict) else None,
            )
        )
    return "\n\n".join(snippets), sources


def _extract_file_text(local_path: Optional[str], file_name: str, mime_type: str) -> str:
    if not local_path or not os.path.exists(local_path):
        return ""
    lower_name = (file_name or "").lower()
    lower_mime = (mime_type or "").lower()
    try:
        if lower_mime.startswith("image/") or lower_name.endswith((".png", ".jpg", ".jpeg", ".webp", ".bmp", ".gif")):
            return _extract_image_text(local_path, mime_type)
        if lower_name.endswith(".txt") or lower_name.endswith(".md") or "text/" in lower_mime:
            return _read_text_file(local_path)
        if lower_name.endswith(".json") or "json" in lower_mime:
            return _read_text_file(local_path)
        if lower_name.endswith(".csv") or "csv" in lower_mime:
            return _read_csv_file(local_path)
        if lower_name.endswith(".xlsx"):
            return _read_xlsx_file(local_path)
        if lower_name.endswith(".xls") or "ms-excel" in lower_mime:
            return _read_xls_file(local_path)
    except Exception:
        return ""
    return ""


def _extract_image_text(local_path: str, mime_type: str) -> str:
    """
    图片文本抽取（主流方案）：
    - 优先走多模态 LLM OCR（无需本地 OCR 依赖）
    - 不可用时自动降级为空文本，不影响上传与入库
    """
    if (
        settings.llm_provider == "mock"
        or not settings.llm_api_base
        or not settings.llm_api_key
        or not local_path
        or not os.path.exists(local_path)
    ):
        return ""

    try:
        with open(local_path, "rb") as f:
            image_bytes = f.read()
    except Exception:
        return ""

    if not image_bytes:
        return ""

    data_url = f"data:{mime_type or 'image/png'};base64,{base64.b64encode(image_bytes).decode('utf-8')}"
    payload = {
        "model": settings.llm_model,
        "messages": [
            {
                "role": "system",
                "content": "你是OCR助手。请提取图片中的可读文字，只返回纯文本，不要解释。",
            },
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": "请提取这张图片中的所有可读文字。"},
                    {"type": "image_url", "image_url": {"url": data_url}},
                ],
            },
        ],
        "temperature": 0,
    }

    req = urllib_request.Request(
        url=f"{settings.llm_api_base.rstrip('/')}/chat/completions",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {settings.llm_api_key}",
        },
        method="POST",
    )

    try:
        with urllib_request.urlopen(req, timeout=min(settings.llm_timeout_seconds, 45)) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            text = (
                data.get("choices", [{}])[0]
                .get("message", {})
                .get("content", "")
            )
            if not text:
                return ""
            return str(text).strip()[:20000]
    except Exception:
        return ""


def _read_text_file(path: str) -> str:
    for encoding in ("utf-8", "gbk", "latin-1"):
        try:
            with open(path, "r", encoding=encoding, errors="ignore") as f:
                return f.read(12000)
        except Exception:
            continue
    return ""


def _read_csv_file(path: str) -> str:
    rows: List[str] = []
    with open(path, "r", encoding="utf-8", errors="ignore", newline="") as f:
        reader = csv.reader(f)
        for idx, row in enumerate(reader):
            if idx >= 200:
                break
            rows.append(" | ".join([str(x) for x in row if str(x).strip()]))
    return "\n".join(rows)


def _read_xlsx_file(path: str) -> str:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        lines: List[str] = []
        for ws in reversed(wb.worksheets):
            lines.append(f"[sheet] {ws.title}")
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if row_idx >= 1000:
                    break
                cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                if cells:
                    lines.append(" | ".join(cells))
                if len(lines) >= _MAX_EXTRACT_LINES:
                    break
            if len(lines) >= _MAX_EXTRACT_LINES:
                break
        normalized = _normalize_sheet_matrix(lines)
        # 优先保留结构化三元组，避免中段被裁剪导致“姓名+日期+事项”丢失
        merged = normalized if normalized else lines
        text = _fit_text_budget("\n".join(merged), 50000)
        if text.strip():
            return text
    except Exception:
        pass
    # 兜底：部分 WPS/非标准 xlsx 会在 openpyxl 样式解析失败，改用 XML 直读
    return _read_xlsx_xml_fallback(path)


def _read_xls_file(path: str) -> str:
    with open(path, "rb") as f:
        content = f.read()
    wb = xlrd.open_workbook(file_contents=content)
    lines: List[str] = []
    for sheet_idx in range(wb.nsheets - 1, -1, -1):
        ws = wb.sheet_by_index(sheet_idx)
        lines.append(f"[sheet] {ws.name}")
        for row_idx in range(min(ws.nrows, 1000)):
            cells = []
            for col_idx in range(ws.ncols):
                val = ws.cell_value(row_idx, col_idx)
                text = str(val).strip()
                if text:
                    cells.append(text)
            if cells:
                lines.append(" | ".join(cells))
            if len(lines) >= _MAX_EXTRACT_LINES:
                break
        if len(lines) >= _MAX_EXTRACT_LINES:
            break
    normalized = _normalize_sheet_matrix(lines)
    merged = normalized if normalized else lines
    return _fit_text_budget("\n".join(merged), 50000)


def _read_xlsx_xml_fallback(path: str) -> str:
    ns = {"s": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    lines: List[str] = []
    matrix_lines: List[str] = []
    try:
        with zipfile.ZipFile(path, "r") as z:
            shared_strings: List[str] = []
            if "xl/sharedStrings.xml" in z.namelist():
                root = ET.fromstring(z.read("xl/sharedStrings.xml"))
                for si in root.findall("s:si", ns):
                    # si 可能由多个 r/t 组成
                    parts = []
                    for t in si.findall(".//s:t", ns):
                        if t.text:
                            parts.append(t.text)
                    shared_strings.append("".join(parts).strip())

            wb_root = ET.fromstring(z.read("xl/workbook.xml"))
            sheet_nodes = wb_root.findall("s:sheets/s:sheet", ns)
            rel_root = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
            rel_ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
            rel_map = {}
            for rel in rel_root.findall("r:Relationship", rel_ns):
                rel_map[rel.attrib.get("Id")] = rel.attrib.get("Target", "")

            for sheet in reversed(sheet_nodes):
                sheet_name = sheet.attrib.get("name", "sheet")
                rid = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
                target = rel_map.get(rid or "", "")
                if not target:
                    continue
                if not target.startswith("worksheets/"):
                    continue
                sheet_xml_path = f"xl/{target}"
                if sheet_xml_path not in z.namelist():
                    continue
                lines.append(f"[sheet] {sheet_name}")
                matrix_lines.append(f"[sheet] {sheet_name}")
                sheet_root = ET.fromstring(z.read(sheet_xml_path))
                rows = sheet_root.findall("s:sheetData/s:row", ns)
                sheet_rows: List[List[str]] = []
                for row_idx, row in enumerate(rows):
                    if row_idx >= 1000:
                        break
                    vals = []
                    for c in row.findall("s:c", ns):
                        cell_type = c.attrib.get("t")
                        v = c.find("s:v", ns)
                        if v is None or v.text is None:
                            continue
                        raw = v.text.strip()
                        if not raw:
                            continue
                        if cell_type == "s":
                            try:
                                si = int(raw)
                                value = shared_strings[si] if 0 <= si < len(shared_strings) else raw
                            except Exception:
                                value = raw
                        else:
                            value = raw
                        value = value.strip()
                        if value:
                            vals.append(value)
                    if vals:
                        lines.append(" | ".join(vals))
                        sheet_rows.append(vals)
                    if len(lines) >= _MAX_EXTRACT_LINES:
                        break
                matrix_lines.extend(_build_member_date_triples(sheet_rows))
                if len(lines) >= _MAX_EXTRACT_LINES:
                    break
    except Exception:
        return ""
    merged = matrix_lines if matrix_lines else lines
    return _fit_text_budget("\n".join(merged), 50000)


def _fit_text_budget(text: str, limit: int) -> str:
    if len(text) <= limit:
        return text
    head = int(limit * 0.65)
    tail = max(limit - head - 32, 0)
    return text[:head] + "\n...[trimmed]...\n" + text[-tail:]


def _is_date_cell(value: str) -> bool:
    v = (value or "").strip()
    if not v:
        return False
    return bool(re.search(r"\d{4}[/-]\d{1,2}[/-]\d{1,2}", v))


def _build_member_date_triples(rows: List[List[str]]) -> List[str]:
    if not rows:
        return []
    header_idx = -1
    for idx, row in enumerate(rows[:20]):
        if any("成员" in c for c in row) and any(_is_date_cell(c) for c in row):
            header_idx = idx
            break
    if header_idx < 0:
        return []

    header = rows[header_idx]
    triples: List[str] = []
    member_counts: Dict[str, int] = {}
    per_member_cap = 8
    for row in rows[header_idx + 1 :]:
        if not row:
            continue
        member = row[0].strip() if len(row) > 0 else ""
        if not member or member in ("成员", "今日工作", "明日工作", "问题"):
            continue
        max_cols = min(len(header), len(row))
        for col_idx in range(1, max_cols):
            date_label = header[col_idx].strip() if col_idx < len(header) else ""
            if not _is_date_cell(date_label):
                continue
            task = row[col_idx].strip() if col_idx < len(row) else ""
            if not task:
                continue
            used = member_counts.get(member, 0)
            if used >= per_member_cap:
                continue
            triples.append(f"{member} | {date_label} | {task}")
            member_counts[member] = used + 1
            if len(triples) >= 2500:
                return triples
    return triples


def _normalize_sheet_matrix(lines: List[str]) -> List[str]:
    normalized: List[str] = []
    current_rows: List[List[str]] = []
    has_structured_rows = False
    for line in lines:
        if not line:
            continue
        if line.startswith("[sheet] "):
            if current_rows:
                triples = _build_member_date_triples(current_rows)
                if triples:
                    has_structured_rows = True
                    normalized.extend(triples)
                current_rows = []
            normalized.append(line)
            continue
        current_rows.append([c.strip() for c in line.split("|") if c.strip()])
    if current_rows:
        triples = _build_member_date_triples(current_rows)
        if triples:
            has_structured_rows = True
            normalized.extend(triples)
    # 仅抽取到 sheet 标题而没有结构化三元组时，回退到原始行，避免内容被误丢弃。
    if not has_structured_rows:
        return []
    return normalized


async def rebuild_uploaded_files_kb(limit: int = 500) -> Dict[str, int]:
    docs = await list_chat_file_documents(limit=limit)
    total = 0
    ok = 0
    skipped = 0
    for doc in docs:
        total += 1
        source_id = str(doc.get("source_id") or "").strip()
        metadata = doc.get("metadata_json")
        if not source_id:
            skipped += 1
            continue
        if isinstance(metadata, str):
            try:
                metadata = json.loads(metadata)
            except Exception:
                metadata = {}
        metadata = metadata or {}
        file_name = str(metadata.get("name") or "附件")
        mime_type = str(metadata.get("mime_type") or "application/octet-stream")
        file_size = int(metadata.get("size") or 0)
        filename = source_id.rstrip("/").split("/")[-1]
        local_path = os.path.join(settings.upload_dir, filename) if filename else ""
        if not local_path or not os.path.exists(local_path):
            skipped += 1
            continue
        await ingest_uploaded_file(
            room_id=str(doc.get("room_id") or settings.default_room),
            file_url=source_id,
            file_name=file_name,
            mime_type=mime_type,
            file_size=file_size,
            local_path=local_path,
        )
        ok += 1
    return {"total": total, "ok": ok, "skipped": skipped}
